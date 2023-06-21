from openpyxl import load_workbook
from pycomm3 import LogixDriver, PycommError
from datetime import datetime, timezone


class PlcHistory:
    """Version Control

    Version Control through GitHub, see link below to repository.
    https://github.com/Site-Automation/plc_tag_history_retrieval.git
    """

    def __init__(self, file_path):
        """Private Variables

        Leading underscores create private class members.

        https://docs.python.org/3/tutorial/classes.html#private-variables
        """
        self._file_path = file_path
        self._work_book = load_workbook(filename=self._file_path)
        self._work_sheet = self._work_book.sheetnames

    def __str__(self):
        return f"Worksheets from the spreadsheet: {self._work_sheets}"

    """Property Decorator
    
    The is the pythonic way and recommendation for using getter and setters.
    Avoid creating and naming functions with the word 'get' and 'set'. See link below.
    
    https://docs.python.org/3.10/library/functions.html?highlight=property#property
    """

    @property  # Create a getter using the property decorator.
    def work_sheet(self):
        return self._work_sheet

    @work_sheet.setter  # Create a setter using the property decorator.
    def work_sheet(self, work_sheet):
        self._work_sheet = work_sheet

    # Return the row with tag names from the Excel spreadsheet and put those names into a list.
    def plc_tag_names(self, plc_tags_start_row=12):
        work_sheet = self._work_book[self.work_sheet]
        plc_tags_list = [tag for row in
                         work_sheet.iter_rows(min_row=plc_tags_start_row, max_row=plc_tags_start_row, min_col=2,
                                              values_only=True) for tag in row]

        return plc_tags_list

    def plc_tag_values(self, plc_tags_name_list):
        """PLC Ethernet/IP Connection

        Open a connection to the specified PLC from the ipv4 address set in the excel spreadsheet.
        If the connection is successful, use the tag name list returned from the plc_tags_names() function and start
        determining what operation to perform on that tag based on if it's a plc tag array of values or epoch times.
        Once the type of plc tag is determined, get each element in the array and parse the value or epoch time.
        Create a nested list (plc_tag_value_list) of values for each tag.
        Return a nested list that is created by unpacking each elements related position in the previous nested list (plc_tag_value_list).
        For example a nested list is created that gives every element[0] in a nested list.

         https://github.com/ottowayi/pycomm3
        """

        ipv4_address = self._work_book[self.work_sheet]['C5'].value
        plc = LogixDriver(ipv4_address)
        plc_tags_value_list = []
        failed_tag_count = 0
        passed_tag_count = 0

        try:
            plc.open()
            for tag in plc_tags_name_list:
                try:
                    if "time" in tag.lower():
                        temp_date_list = [datetime.fromtimestamp(time, tz=timezone.utc).strftime("%d-%b-%Y %H:%M:%S")
                                          for time in plc.read(tag + "{2000}").value]
                        plc_tags_value_list.append(temp_date_list)
                    else:
                        temp_value_list = [round(value, 4) for value in plc.read(tag + "{2000}").value]
                        plc_tags_value_list.append(temp_value_list)
                    passed_tag_count += 1
                except TypeError as e:
                    print(f"Plc tag '{tag}' doesn't exist!!!\n")
                    failed_tag_count += 1
                    plc_tags_value_list.append(["Error" for _ in range(0, 2000)])
            print(f"{failed_tag_count} of {len(plc_tags_name_list)} tags failed during processing!")
            print(f"{passed_tag_count} of {len(plc_tags_name_list)} tags processed successfully!")
        except PycommError as e:
            print(e)

        return list(map(list, zip(*plc_tags_value_list)))

    # Take the list returned from the plc_tag_values() loop through each nested lest and write each element into a cell
    # in a Excel spreadsheet.
    def write_tag_values_wb(self, tags_value_list):
        work_book = self._work_book
        work_sheet = self._work_book[self.work_sheet]
        for i, row in enumerate(tags_value_list):
            for k, value in enumerate(row):
                work_sheet.cell(row=i + 13, column=k + 2).value = value
        work_book.save(filename=self._file_path)

    # Take the start and end time it took to execute the program and find the elapsed time then write the start, end,
    # and elapsed time into a Excel spreadsheet.
    def write_exec_time_wb(self, start_time, end_time):
        work_book = self._work_book
        work_sheet = self._work_book[self.work_sheet]
        elapsed_time = end_time - start_time
        work_sheet.cell(row=8, column=3).value = start_time.strftime("%d-%b-%Y %H:%M:%S")
        work_sheet.cell(row=9,
                        column=3).value = f"Completed {end_time.strftime('%d-%b-%Y %H:%M:%S')} - elapsed time {elapsed_time} seconds"
        work_book.save(filename=self._file_path)
        print(f"Ran program in {elapsed_time} seconds.")
